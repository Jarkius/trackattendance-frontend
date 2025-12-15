"""Dashboard service for multi-station scan reports (Issue #27).

This module provides centralized reporting by:
- Querying Cloud API for multi-station scan data
- Reading local SQLite for employee count (registered headcount)

The dashboard shows:
- Total registered employees (from local SQLite)
- Total unique badges scanned (from Cloud API - all stations)
- Attendance rate (scanned / registered)
- Per-station breakdown
- Excel export functionality
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

import requests

if TYPE_CHECKING:
    from database import DatabaseManager

logger = logging.getLogger(__name__)


class DashboardService:
    """Service for fetching multi-station dashboard data via Cloud API."""

    def __init__(
        self,
        db_manager: "DatabaseManager",
        api_url: str,
        api_key: str,
    ) -> None:
        """Initialize the dashboard service.

        Args:
            db_manager: Local SQLite database manager (for employee count)
            api_url: Cloud API base URL
            api_key: Cloud API authentication key
        """
        self._db_manager = db_manager
        self._api_url = api_url.rstrip("/")
        self._api_key = api_key
        self._timeout = 15  # seconds

    def _get_headers(self) -> Dict[str, str]:
        """Get HTTP headers for API requests."""
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/json",
        }

    def get_dashboard_data(self) -> Dict[str, Any]:
        """Fetch all dashboard data.

        Returns:
            Dictionary containing:
            - registered: int (employee count from local SQLite)
            - scanned: int (unique badges from cloud)
            - total_scans: int (total scans from cloud)
            - attendance_rate: float (percentage)
            - stations: list of station data
            - last_updated: str (ISO timestamp)
            - error: str (if any error occurred)
        """
        result = {
            "registered": 0,
            "scanned": 0,
            "total_scans": 0,
            "attendance_rate": 0.0,
            "stations": [],
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error": None,
        }

        # Get registered employee count from local SQLite
        try:
            result["registered"] = self._db_manager.count_employees()
            logger.debug(f"Dashboard: Local employee count = {result['registered']}")
        except Exception as e:
            logger.error(f"Dashboard: Failed to get employee count: {e}")
            result["error"] = f"Failed to get employee count: {e}"

        # Get cloud scan data from API
        try:
            response = requests.get(
                f"{self._api_url}/v1/dashboard/stats",
                headers=self._get_headers(),
                timeout=self._timeout,
            )

            if response.status_code == 200:
                data = response.json()
                result["total_scans"] = data.get("total_scans", 0)
                result["scanned"] = data.get("unique_badges", 0)
                result["stations"] = [
                    {
                        "name": s.get("name", "--"),
                        "scans": s.get("scans", 0),
                        "unique": s.get("unique", 0),
                        "last_scan": self._format_time(s.get("last_scan")),
                    }
                    for s in data.get("stations", [])
                ]
                logger.info(
                    f"Dashboard: total_scans={result['total_scans']}, "
                    f"unique_badges={result['scanned']}, stations={len(result['stations'])}"
                )
            elif response.status_code == 401:
                result["error"] = "Authentication failed - check API key"
                logger.error("Dashboard: API authentication failed")
            elif response.status_code == 503:
                result["error"] = "Cloud database unavailable"
                logger.error("Dashboard: Cloud database unavailable")
            else:
                result["error"] = f"API error: {response.status_code}"
                logger.error(f"Dashboard: API returned {response.status_code}")

        except requests.exceptions.ConnectionError:
            result["error"] = "Cannot connect to cloud API"
            logger.error("Dashboard: Connection error")
        except requests.exceptions.Timeout:
            result["error"] = "Cloud API timeout"
            logger.error("Dashboard: Request timeout")
        except Exception as e:
            result["error"] = f"API error: {e}"
            logger.error(f"Dashboard: Unexpected error: {e}")

        # Calculate attendance rate
        if result["registered"] > 0:
            result["attendance_rate"] = round(
                (result["scanned"] / result["registered"]) * 100, 1
            )

        return result

    def _format_time(self, iso_timestamp: Optional[str]) -> str:
        """Format ISO timestamp to time-only string."""
        if not iso_timestamp:
            return "--"
        try:
            dt = datetime.fromisoformat(iso_timestamp.replace("Z", "+00:00"))
            return dt.strftime("%H:%M:%S")
        except Exception:
            return "--"

    def export_to_excel(self, file_path: str) -> Dict[str, Any]:
        """Export dashboard data to Excel file.

        Args:
            file_path: Path where Excel file will be saved

        Returns:
            Dictionary with:
            - ok: bool (success status)
            - message: str (success/error message)
            - file_path: str (path to created file)
        """
        result = {
            "ok": False,
            "message": "",
            "file_path": file_path,
        }

        # Fetch export data from API
        try:
            response = requests.get(
                f"{self._api_url}/v1/dashboard/export",
                headers=self._get_headers(),
                timeout=60,  # Longer timeout for export
            )

            if response.status_code != 200:
                result["message"] = f"API error: {response.status_code}"
                return result

            data = response.json()
            scans = data.get("scans", [])

            if not scans:
                result["message"] = "No scan data to export"
                return result

        except requests.exceptions.ConnectionError:
            result["message"] = "Cannot connect to cloud API"
            return result
        except requests.exceptions.Timeout:
            result["message"] = "Cloud API timeout"
            return result
        except Exception as e:
            result["message"] = f"API error: {e}"
            return result

        # Generate Excel file
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            # Create DataFrame from API data
            df = pd.DataFrame(scans)
            df.columns = ["Badge ID", "Station", "Scanned At", "Matched"]

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "All Scans"

            # Add header with styling
            header_fill = PatternFill(start_color="86bc25", end_color="86bc25", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            columns = ["Badge ID", "Station", "Scanned At", "Matched"]
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Add data rows
            for row_idx, row in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

            # Add summary sheet
            ws_summary = wb.create_sheet("Summary")
            dashboard_data = self.get_dashboard_data()

            ws_summary["A1"] = "Metric"
            ws_summary["B1"] = "Value"
            ws_summary["A1"].font = header_font
            ws_summary["A1"].fill = header_fill
            ws_summary["B1"].font = header_font
            ws_summary["B1"].fill = header_fill

            summary_data = [
                ("Registered Employees", dashboard_data["registered"]),
                ("Unique Badges Scanned", dashboard_data["scanned"]),
                ("Total Scans", dashboard_data["total_scans"]),
                ("Attendance Rate", f"{dashboard_data['attendance_rate']}%"),
                ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]

            for row_idx, (label, value) in enumerate(summary_data, start=2):
                ws_summary.cell(row=row_idx, column=1, value=label)
                ws_summary.cell(row=row_idx, column=2, value=value)

            ws_summary.column_dimensions["A"].width = 25
            ws_summary.column_dimensions["B"].width = 20

            # Add station breakdown sheet
            if dashboard_data["stations"]:
                ws_stations = wb.create_sheet("By Station")
                station_headers = ["Station", "Total Scans", "Unique Badges", "Last Scan"]

                for col_idx, col_name in enumerate(station_headers, start=1):
                    cell = ws_stations.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_idx, station in enumerate(dashboard_data["stations"], start=2):
                    ws_stations.cell(row=row_idx, column=1, value=station["name"])
                    ws_stations.cell(row=row_idx, column=2, value=station["scans"])
                    ws_stations.cell(row=row_idx, column=3, value=station["unique"])
                    ws_stations.cell(row=row_idx, column=4, value=station["last_scan"])

                for col in ws_stations.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws_stations.column_dimensions[col[0].column_letter].width = max_length + 2

            # Save file
            wb.save(file_path)
            result["ok"] = True
            result["message"] = f"Exported {len(df)} scans to Excel"
            logger.info(f"Dashboard: Exported to {file_path}")

        except ImportError as e:
            result["message"] = f"Missing dependency: {e}. Run: pip install pandas openpyxl"
            logger.error(result["message"])
        except Exception as e:
            result["message"] = f"Export failed: {e}"
            logger.error(f"Dashboard export error: {e}")

        return result


__all__ = ["DashboardService"]
