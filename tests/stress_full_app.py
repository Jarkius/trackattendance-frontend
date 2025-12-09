import argparse
import itertools
import json
import random
import sys
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from PyQt6.QtCore import QEventLoop, QTimer

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from attendance import AttendanceService
from sync import SyncService
import config
from main import (
    Api,
    DATABASE_PATH,
    EMPLOYEE_WORKBOOK_PATH,
    EXPORT_DIRECTORY,
    initialize_app,
)

from openpyxl import load_workbook

SPECIAL_CASE_BARCODES = [
    '999999',               # intentionally invalid control scan
    '!@#$%',                # punctuation-only input
    '12345-ABC',            # mixed digits and hyphen
    'DROP TABLE;',          # SQL-ish input
    '"quoted"',            # quotes inside payload
    "');--",               # closing quote + comment
]


def _cycle_barcodes(iterations: int, base: Iterable[str]) -> List[str]:
    base_list = list(base)
    if not base_list:
        return []
    iterator = itertools.cycle(base_list)
    return [next(iterator) for _ in range(iterations)]


def _load_employee_barcodes(workbook_path: Path) -> List[str]:
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {
            str(name).strip(): idx
            for idx, name in enumerate(header_row)
            if name and str(name).strip()
        }
        legacy_index = header_map.get('Legacy ID')
        if legacy_index is None:
            return []
        seen = set()
        barcodes: List[str] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cell = row[legacy_index]
            if cell is None:
                continue
            value = str(cell).strip()
            if not value or value in seen:
                continue
            seen.add(value)
            barcodes.append(value)
        return barcodes
    finally:
        workbook.close()


def _sample_employee_barcodes(workbook_path: Path, sample_size: int) -> List[str]:
    employees = _load_employee_barcodes(workbook_path)
    if not employees:
        return []
    sample_size = max(1, min(sample_size, len(employees)))
    if sample_size == len(employees):
        return employees
    return random.sample(employees, sample_size)


def _default_barcodes(sample_size: int, include_specials: bool = True) -> List[str]:
    employee_samples = _sample_employee_barcodes(EMPLOYEE_WORKBOOK_PATH, sample_size)
    barcodes = employee_samples
    if include_specials:
        barcodes = barcodes + SPECIAL_CASE_BARCODES
    if not barcodes:
        return SPECIAL_CASE_BARCODES if include_specials else []
    return barcodes


def _ensure_station_name(service: AttendanceService, fallback: str = 'Stress Harness') -> str:
    try:
        station = service.station_name
        if station:
            return station
    except RuntimeError:
        pass

    name = fallback.strip() or 'Station'
    service._db.set_station_name(name)  # type: ignore[attr-defined]
    service._station_name = name  # type: ignore[attr-defined]
    return name


def _run_js(view, script: str) -> Any:
    loop = QEventLoop()
    result_container: Dict[str, Any] = {}

    def handle_result(result: Any) -> None:
        result_container['value'] = result
        loop.quit()

    view.page().runJavaScript(script, handle_result)
    loop.exec()
    return result_container.get('value')


def _dispatch_scan(view, barcode: str) -> Dict[str, Any]:
    payload = json.dumps(barcode)
    script = f"""
        (function(barcode) {{
            const barcodeInput = document.getElementById('barcode-input');
            const feedback = document.getElementById('live-feedback-name');
            const totalScanned = document.getElementById('total-scanned');
            const historyList = document.getElementById('scan-history-list');
            if (!barcodeInput) {{
                return {{ status: 'missing-input', barcode }};
            }}
            barcodeInput.value = barcode;
            const event = new KeyboardEvent('keyup', {{ key: 'Enter', bubbles: true }});
            barcodeInput.dispatchEvent(event);
            const firstHistory = historyList && historyList.firstElementChild;
            const historyName = firstHistory ? firstHistory.querySelector('.name') : null;
            return {{
                status: 'ok',
                barcode,
                feedbackText: feedback ? feedback.textContent.trim() : null,
                feedbackColor: feedback ? window.getComputedStyle(feedback).color : null,
                totalScanned: totalScanned ? totalScanned.textContent.trim() : null,
                historyTop: historyName ? historyName.textContent.trim() : null
            }};
        }})({payload});
    """
    result = _run_js(view, script)
    if not isinstance(result, dict):
        return {'status': 'no-result', 'barcode': barcode}
    if 'barcode' not in result:
        result['barcode'] = barcode
    return result


def _collect_snapshot(view) -> Dict[str, Any]:
    script = """
        (function() {
            const feedback = document.getElementById('live-feedback-name');
            const totalScanned = document.getElementById('total-scanned');
            const historyList = document.getElementById('scan-history-list');
            const historyName = historyList && historyList.firstElementChild ? historyList.firstElementChild.querySelector('.name') : null;
            return {
                feedbackText: feedback ? feedback.textContent.trim() : null,
                feedbackColor: feedback ? window.getComputedStyle(feedback).color : null,
                totalScanned: totalScanned ? totalScanned.textContent.trim() : null,
                historyTop: historyName ? historyName.textContent.trim() : null
            };
        })();
    """
    snapshot = _run_js(view, script)
    return snapshot if isinstance(snapshot, dict) else {}


def run_stress_test(
    iterations: int,
    barcodes: Sequence[str],
    delay_ms: int,
    show_window: bool,
    show_full_screen: bool,
    enable_fade: bool,
    verbose: bool,
) -> int:
    load_state: Dict[str, Any] = {'ok': None}
    load_loop = QEventLoop()

    service = AttendanceService(
        database_path=DATABASE_PATH,
        employee_workbook_path=EMPLOYEE_WORKBOOK_PATH,
        export_directory=EXPORT_DIRECTORY,
    )
    _ensure_station_name(service)

    # Initialize sync service for cloud integration testing
    sync_service = SyncService(
        db=service._db,
        api_url=config.CLOUD_API_URL,
        api_key=config.CLOUD_API_KEY,
        batch_size=config.CLOUD_SYNC_BATCH_SIZE,
    )

    def on_load_finished(ok: bool) -> None:
        load_state['ok'] = ok
        if load_loop.isRunning():
            load_loop.quit()

    def api_factory(quit_callback):
        return Api(service=service, quit_callback=quit_callback)

    try:
        app, window, view, _ = initialize_app(
            argv=sys.argv,
            show_window=show_window,
            show_full_screen=show_full_screen,
            enable_fade=enable_fade,
            on_load_finished=on_load_finished,
            api_factory=api_factory,
        )

        window.setProperty('suppress_export_notification', True)

        if load_state['ok'] is None:
            load_loop.exec()
        if not load_state['ok']:
            return 3

        sequence = _cycle_barcodes(iterations, barcodes)
        failures: List[Dict[str, Any]] = []
        successful = 0
        invalid = 0
        start = time.perf_counter()

        for index, barcode in enumerate(sequence, start=1):
            result = _dispatch_scan(view, barcode)
            status = result.get('status', 'unknown')
            feedback = (result.get('feedbackText') or '').strip()
            total_scanned = result.get('totalScanned')
            history_top = result.get('historyTop')

            if status != 'ok':
                failures.append(result)
            else:
                if feedback.lower().startswith('not matched'):
                    invalid += 1
                else:
                    successful += 1

            if verbose:
                print(f"[{index:03d}] status={status} barcode={barcode} feedback='{feedback}' total={total_scanned} history='{history_top}'")
            elif index % 25 == 0 or index == 1 or index == iterations:
                print(f"[{index:03d}/{iterations}] status={status} feedback='{feedback}' total={total_scanned}")

            if delay_ms > 0:
                settle_loop = QEventLoop()
                QTimer.singleShot(delay_ms, settle_loop.quit)
                settle_loop.exec()

        duration = time.perf_counter() - start
        snapshot = _collect_snapshot(view)

        # Test sync service before export (simulates shutdown sync)
        sync_attempted = False
        sync_success = False
        synced_count = 0
        failed_count = 0

        if sync_service:
            try:
                stats_before = sync_service.db.get_sync_statistics()
                pending_before = stats_before.get('pending', 0)

                if pending_before > 0:
                    sync_attempted = True
                    print(f'[sync] Syncing {pending_before} pending scan(s) before export...')
                    sync_start = time.perf_counter()
                    sync_result = sync_service.sync_pending_scans()
                    sync_duration = time.perf_counter() - sync_start

                    synced_count = sync_result.get('synced', 0)
                    failed_count = sync_result.get('failed', 0)
                    pending_after = sync_result.get('pending', 0)

                    sync_success = synced_count > 0 or failed_count == 0

                    print(f'[sync] Complete in {sync_duration:.2f}s: {synced_count} synced, {failed_count} failed, {pending_after} pending')
                else:
                    print('[sync] No pending scans to sync')
            except Exception as exc:
                print(f'[sync] Sync failed: {exc}')
                sync_attempted = True
                sync_success = False

        export_info = None
        try:
            export_info = service.export_scans()
        except Exception as exc:
            if verbose:
                print(f'[warn] export failed: {exc}')
        else:
            if export_info and verbose:
                dest = export_info.get('absolutePath') or export_info.get('fileName')
                if dest:
                    print(f'[info] export written to {dest}')
            window.setProperty('export_notification_triggered', True)

        window.close()
        for _ in range(3):
            app.processEvents()

        print("\n--- Stress Test Summary ---")
        print(f'Scans attempted : {iterations}')
        print(f'Successful scans: {successful}')
        print(f'Invalid scans   : {invalid}')
        print(f'Failures        : {len(failures)}')
        print(f'Total runtime   : {duration:.2f}s')

        if sync_attempted:
            print(f'\n--- Cloud Sync Results ---')
            print(f'Sync attempted  : Yes')
            print(f'Sync success    : {"Yes" if sync_success else "No"}')
            print(f'Scans synced    : {synced_count}')
            print(f'Scans failed    : {failed_count}')
        else:
            print(f'\n--- Cloud Sync Results ---')
            print(f'Sync attempted  : No (no pending scans)')

        if failures:
            print("\nFirst failure:")
            print(json.dumps(failures[0], indent=2))
            return 2

        expected_total = str(successful)
        final_total = snapshot.get('totalScanned') if snapshot else None
        if final_total and final_total != expected_total:
            print(f"Warning: total_scanned reported as {final_total}, expected {expected_total}.")

        return 0
    finally:
        service.close()


def main() -> int:
    parser = argparse.ArgumentParser(description='Drive the full PyQt window and simulate barcode scans.')
    parser.add_argument('barcodes', nargs='*', help='Base barcode values to cycle through; defaults use random workbook samples.')
    parser.add_argument('--iterations', type=int, default=200, help='Number of barcode submissions to perform.')
    parser.add_argument('--delay-ms', type=int, default=75, help='Delay between scans to mimic hardware pacing.')
    parser.add_argument('--sample-size', type=int, default=50, help='Employee barcodes to sample from the workbook when no explicit list is provided.')
    parser.add_argument('--no-specials', action='store_true', help='Exclude synthetic invalid barcode cases from the run.')
    parser.add_argument('--no-show-window', action='store_true', help='Keep the window hidden during the run.')
    parser.add_argument('--windowed', action='store_true', help='Show the window but avoid fullscreen mode.')
    parser.add_argument('--disable-fade', action='store_true', help='Skip the window fade animation to save a few frames.')
    parser.add_argument('--verbose', action='store_true', help='Log every scan instead of periodic checkpoints.')
    args = parser.parse_args()

    if args.barcodes:
        base_barcodes = list(args.barcodes)
    else:
        base_barcodes = _default_barcodes(sample_size=max(args.sample_size, 1), include_specials=not args.no_specials)

    if not base_barcodes:
        print('No barcodes available for the stress test.', file=sys.stderr)
        return 1

    status = run_stress_test(
        iterations=args.iterations,
        barcodes=base_barcodes,
        delay_ms=max(args.delay_ms, 0),
        show_window=not args.no_show_window,
        show_full_screen=not args.windowed,
        enable_fade=not args.disable_fade,
        verbose=args.verbose,
    )
    return status


if __name__ == '__main__':
    sys.exit(main())
