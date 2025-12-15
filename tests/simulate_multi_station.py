"""Simulate scan data from multiple stations for dashboard testing.

This script sends simulated scans from multiple stations to the Cloud API
to test the dashboard's ability to display 10+ stations and BU breakdown.

Usage:
    python tests/simulate_multi_station.py --stations 10 --scans-per-station 20
"""

import argparse
import random
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import requests
from openpyxl import load_workbook
import config

STATION_NAMES = [
    "Main Gate",
    "Side Entrance",
    "North Wing",
    "South Wing",
    "Tower A Lobby",
    "Tower B Lobby",
    "Parking Level 1",
    "Parking Level 2",
    "VIP Entrance",
    "Staff Entrance",
    "Basement Access",
    "Rooftop Helipad",
]


def load_employee_badges(workbook_path: Path, sample_size: int = 50) -> list[str]:
    """Load employee badge IDs from the workbook."""
    if not workbook_path.exists():
        print(f"[ERROR] Employee workbook not found: {workbook_path}")
        return []

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {
            str(name).strip(): idx
            for idx, name in enumerate(header_row)
            if name and str(name).strip()
        }
        legacy_index = header_map.get("Legacy ID")
        if legacy_index is None:
            return []

        badges = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cell = row[legacy_index]
            if cell:
                badges.append(str(cell).strip())

        if len(badges) > sample_size:
            badges = random.sample(badges, sample_size)
        return badges
    finally:
        workbook.close()


def send_scans_to_api(station_name: str, badge_ids: list[str]) -> dict:
    """Send scan events to the Cloud API."""
    events = []
    base_time = datetime.now(timezone.utc)

    for i, badge_id in enumerate(badge_ids):
        # Generate unique idempotency key
        idempotency_key = f"{station_name.replace(' ', '-')}-{badge_id}-test-{base_time.strftime('%Y%m%d%H%M%S')}-{i}"

        events.append({
            "idempotency_key": idempotency_key,
            "badge_id": badge_id,
            "station_name": station_name,
            "scanned_at": base_time.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "meta": {
                "matched": True,
                "local_id": 999000 + i,
            }
        })

    headers = {
        "Authorization": f"Bearer {config.CLOUD_API_KEY}",
        "Content-Type": "application/json; charset=utf-8",
    }

    response = requests.post(
        f"{config.CLOUD_API_URL}/v1/scans/batch",
        headers=headers,
        json={"events": events},
        timeout=30,
    )

    if response.status_code == 200:
        return response.json()
    else:
        return {"error": f"HTTP {response.status_code}: {response.text}"}


def get_dashboard_stats() -> dict:
    """Fetch current dashboard stats."""
    headers = {
        "Authorization": f"Bearer {config.CLOUD_API_KEY}",
        "Content-Type": "application/json",
    }

    response = requests.get(
        f"{config.CLOUD_API_URL}/v1/dashboard/stats",
        headers=headers,
        timeout=15,
    )

    if response.status_code == 200:
        return response.json()
    return {}


def main():
    parser = argparse.ArgumentParser(description="Simulate multi-station scans for dashboard testing")
    parser.add_argument("--stations", type=int, default=10, help="Number of stations to simulate")
    parser.add_argument("--scans-per-station", type=int, default=15, help="Scans per station")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be sent without sending")
    args = parser.parse_args()

    # Load employee badges
    workbook_path = ROOT_DIR / "data" / "employee.xlsx"
    badges = load_employee_badges(workbook_path, sample_size=100)

    if not badges:
        print("[ERROR] No employee badges found")
        return 1

    print(f"Loaded {len(badges)} employee badges")

    # Select stations
    stations = STATION_NAMES[:args.stations]
    print(f"Simulating {len(stations)} stations with {args.scans_per_station} scans each")
    print(f"Stations: {', '.join(stations)}")

    if args.dry_run:
        print("\n[DRY RUN] Would send:")
        for station in stations:
            sample_badges = random.sample(badges, min(args.scans_per_station, len(badges)))
            print(f"  {station}: {len(sample_badges)} scans")
        return 0

    # Send scans for each station
    total_saved = 0
    total_duplicates = 0

    for station in stations:
        sample_badges = random.sample(badges, min(args.scans_per_station, len(badges)))
        print(f"\n[{station}] Sending {len(sample_badges)} scans...")

        result = send_scans_to_api(station, sample_badges)

        if "error" in result:
            print(f"  [ERROR] {result['error']}")
        else:
            saved = result.get("saved", 0)
            duplicates = result.get("duplicates", 0)
            total_saved += saved
            total_duplicates += duplicates
            print(f"  [OK] Saved: {saved}, Duplicates: {duplicates}")

    # Show dashboard stats after upload
    print("\n" + "=" * 50)
    print("DASHBOARD STATS AFTER UPLOAD")
    print("=" * 50)

    stats = get_dashboard_stats()
    print(f"Total scans: {stats.get('total_scans', 0)}")
    print(f"Unique badges: {stats.get('unique_badges', 0)}")
    print(f"\nStations ({len(stats.get('stations', []))}):")
    for s in stats.get("stations", []):
        print(f"  {s.get('name', '--')}: {s.get('scans', 0)} scans, {s.get('unique', 0)} unique")

    print(f"\n--- Summary ---")
    print(f"Total saved: {total_saved}")
    print(f"Total duplicates: {total_duplicates}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
