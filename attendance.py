"""Core attendance application services."""

from __future__ import annotations

import hashlib
import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QInputDialog, QMessageBox, QWidget

from database import DatabaseManager, EmployeeRecord, ScanRecord, ISO_TIMESTAMP_FORMAT

LOGGER = logging.getLogger(__name__)
REQUIRED_COLUMNS = ["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"]
EXAMPLE_WORKBOOK_NAME = "exampleof_employee.xlsx"


DISPLAY_TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M:%S"
class AttendanceService:
    """High-level operations for coordinating scans, employees, and exports."""

    def __init__(
        self,
        *,
        database_path: Path,
        employee_workbook_path: Path,
        export_directory: Path,
    ) -> None:
        self._db = DatabaseManager(database_path)
        self._employee_workbook_path = employee_workbook_path
        self._example_employee_workbook_path = self._employee_workbook_path.with_name(EXAMPLE_WORKBOOK_NAME)
        self._export_directory = export_directory
        self._export_directory.mkdir(parents=True, exist_ok=True)
        self._employee_headers: List[str] = list(REQUIRED_COLUMNS)
        self._employee_cache: Dict[str, EmployeeRecord] = {}
        self._station_name: Optional[str] = self._db.get_station_name()

        self._bootstrap_employee_directory()
        self._employee_cache = self._db.load_employee_cache()

    def employees_loaded(self) -> bool:
        return self._db.employees_loaded()

    def validate_roster_headers(self, workbook_path: Path) -> tuple[bool, str]:
        """
        Validate that an employee workbook has required columns.

        Returns:
            (is_valid, error_message)
        """
        if not workbook_path.exists():
            return False, f"Roster file not found: {workbook_path.name}"

        try:
            workbook = load_workbook(workbook_path, read_only=True)
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            workbook.close()

            # Check for header row
            if not header_row or all(cell is None for cell in header_row):
                return False, "Roster file has no headers (first row is empty)"

            # Extract non-empty headers
            actual_headers = {
                str(name).strip() for name in header_row
                if name and str(name).strip()
            }

            # Check for required columns
            from config import REQUIRED_ROSTER_COLUMNS
            missing = [col for col in REQUIRED_ROSTER_COLUMNS if col not in actual_headers]

            if missing:
                msg = f"Roster missing required columns:\n\n"
                msg += f"Missing: {', '.join(missing)}\n\n"
                msg += f"Required: {', '.join(REQUIRED_ROSTER_COLUMNS)}\n\n"
                msg += f"Found: {', '.join(sorted(actual_headers)) if actual_headers else '(none)'}"
                return False, msg

            return True, "Roster headers valid"

        except Exception as e:
            return False, f"Error reading roster file: {str(e)}"

    def _bootstrap_employee_directory(self) -> None:
        if not self._employee_workbook_path.exists():
            LOGGER.warning("Employee workbook not found at %s", self._employee_workbook_path)
            self.ensure_example_employee_workbook()
            return

        # Fast path: skip SHA256 if file modification time hasn't changed
        current_mtime = str(self._employee_workbook_path.stat().st_mtime)
        stored_mtime = self._db.get_roster_meta("file_mtime")
        stored_hash = self._db.get_roster_hash()

        if stored_mtime == current_mtime and stored_hash and self._db.employees_loaded():
            LOGGER.info("Roster unchanged (mtime match), skipping reimport")
            return

        # Mtime changed (or first run) — compute hash to confirm
        current_hash = self._hash_file(self._employee_workbook_path)

        if stored_hash == current_hash and self._db.employees_loaded():
            # File was touched but content unchanged — update mtime cache
            self._db.set_roster_meta("file_mtime", current_mtime)
            LOGGER.info("Roster unchanged (hash match, mtime updated), skipping reimport")
            return

        if stored_hash and stored_hash != current_hash:
            LOGGER.info("Roster file changed (hash mismatch), reimporting employees")

        # Validate roster headers before import
        from config import ROSTER_VALIDATION_ENABLED, ROSTER_STRICT_VALIDATION
        is_valid, validation_msg = self.validate_roster_headers(self._employee_workbook_path)

        if not is_valid:
            if ROSTER_VALIDATION_ENABLED:
                LOGGER.error("Roster validation failed: %s", validation_msg)
                if ROSTER_STRICT_VALIDATION:
                    LOGGER.error("Strict validation enabled - skipping import")
                    return
            else:
                LOGGER.warning("Roster validation skipped (disabled): %s", validation_msg)

        workbook = load_workbook(self._employee_workbook_path, read_only=True)
        try:
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            ordered_headers = [
                str(name).strip()
                for name in header_row
                if name and str(name).strip()
            ]
            if ordered_headers:
                self._employee_headers = ordered_headers
            header_to_index = {name: idx for idx, name in enumerate(header_row) if name}
            missing = [name for name in REQUIRED_COLUMNS if name not in header_to_index]
            if missing:
                LOGGER.error("Employee workbook missing columns: %s", ", ".join(missing))
                return

            # Detect optional columns present in this workbook
            email_index = header_to_index.get("Email")

            seen_ids: set[str] = set()
            employees: List[EmployeeRecord] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                legacy_id_raw = row[header_to_index["Legacy ID"]]
                if legacy_id_raw is None:
                    continue
                legacy_id = str(legacy_id_raw).strip()
                if not legacy_id or legacy_id in seen_ids:
                    continue
                full_name = _safe_string(row[header_to_index["Full Name"]])
                sl_l1_desc = _safe_string(row[header_to_index["SL L1 Desc"]])
                position_desc = _safe_string(row[header_to_index["Position Desc"]])
                email = _safe_string(row[email_index]) if email_index is not None else ""
                employees.append(
                    EmployeeRecord(
                        legacy_id=legacy_id,
                        full_name=full_name,
                        sl_l1_desc=sl_l1_desc,
                        position_desc=position_desc,
                        email=email,
                    )
                )
                seen_ids.add(legacy_id)
            if employees:
                # Clear old employees and reimport
                self._db.clear_employees()
                inserted = self._db.bulk_insert_employees(employees)
                self._db.set_roster_hash(current_hash)
                self._db.set_roster_meta("file_mtime", current_mtime)
                LOGGER.info("Imported %s employees from workbook (hash: %s)", inserted, current_hash[:12])
                # Roster BU counts will be pushed to cloud after first
                # successful health check (see main.py Api._run_check)
        finally:
            workbook.close()

    @staticmethod
    def _hash_file(path: Path) -> str:
        """Compute SHA256 hash of a file."""
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()

    def ensure_example_employee_workbook(self) -> Path:
        """Ensure a sample employee roster workbook exists for onboarding."""
        path = self._example_employee_workbook_path
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists():
            return path

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Employees"
        sheet.append(REQUIRED_COLUMNS + ["Email"])

        sample_rows = [
            ("100001", "Ada Lovelace", "Consulting", "Analyst", "alovelace@example.com"),
            ("100002", "Grace Hopper", "Technology", "Engineer", "ghopper@example.com"),
        ]
        for row in sample_rows:
            sheet.append(row)

        workbook.save(path)
        workbook.close()
        return path


    def ensure_station_configured(self, parent: Optional[QWidget] = None) -> str:
        station = self._station_name or self._db.get_station_name()
        if station:
            self._station_name = station
            return station
        while True:
            name, ok = QInputDialog.getText(
                parent,
                "Station Setup",
                "Enter the station name:",
            )
            if not ok:
                QMessageBox.information(
                    parent,
                    "Station Required",
                    "A station name is required to continue. The application will now close.",
                )
                if parent is not None:
                    parent.close()
                self._db.close()
                sys.exit(0)
            sanitized = name.strip()
            if not sanitized:
                QMessageBox.warning(parent, "Invalid Name", "Please provide a non-empty station name.")
                continue
            if len(sanitized) > 50:
                QMessageBox.warning(parent, "Invalid Name", "Station name must be 50 characters or fewer.")
                continue
            if not re.match(r'^[A-Za-z0-9 _-]+$', sanitized):
                QMessageBox.warning(parent, "Invalid Name", "Station name can only contain letters, numbers, spaces, hyphens, and underscores.")
                continue
            self._db.set_station_name(sanitized)
            self._station_name = sanitized
            return sanitized

    @property
    def station_name(self) -> str:
        if self._station_name is None:
            station = self._db.get_station_name()
            if station is None:
                raise RuntimeError("Station name not configured")
            self._station_name = station
        return self._station_name

    def get_initial_payload(self) -> Dict[str, object]:
        import config
        import os

        history = self._db.get_recent_scans()
        return {
            "stationName": self.station_name,
            "totalEmployees": self._db.count_employees(),
            "totalScansToday": self._db.count_scans_today(),
            "totalScansOverall": self._db.count_scans_total(),
            "scanHistory": [_scan_to_dict(scan) for scan in history],
            "connectionCheckIntervalMs": max(0, int(config.CONNECTION_CHECK_INTERVAL_MS)),
            "connectionCheckInitialDelayMs": max(0, int(config.CONNECTION_CHECK_INITIAL_DELAY_MS)),
            "duplicateBadgeAlertDurationMs": max(0, int(config.DUPLICATE_BADGE_ALERT_DURATION_MS)),
            "scanFeedbackDurationMs": max(0, int(config.SCAN_FEEDBACK_DURATION_MS)),
            "debugMode": os.getenv("DEBUG", "False").lower() == "true",
        }

    def search_employee(self, query: str) -> List[Dict[str, object]]:
        """Search employees by email prefix or partial name match."""
        query = query.strip().lower()
        if not query:
            return []
        results = []
        for emp in self._employee_cache.values():
            email_prefix = emp.email.split("@")[0].lower() if emp.email else ""
            name_lower = emp.full_name.lower()
            if (email_prefix and email_prefix == query) or query in name_lower:
                results.append({
                    "legacy_id": emp.legacy_id,
                    "full_name": emp.full_name,
                    "email": emp.email,
                    "business_unit": emp.sl_l1_desc,
                })
            if len(results) >= 10:
                break
        return results

    def register_scan(self, badge_id: str, scan_source: str = "badge",
                       lookup_legacy_id: str = None) -> Dict[str, object]:
        import re
        import config

        sanitized = badge_id.strip()
        if not sanitized:
            return {
                "ok": False,
                "message": "Badge ID is required.",
            }

        # Only derive scan_source when caller used default (submit_scan passes "badge")
        # Lookup/manual paths pass explicit scan_source — don't override
        if scan_source == "badge":
            scan_source = "badge" if re.match(r'^\d+[A-Za-z]?$', sanitized) else "manual"

        # For lookup: user typed a name but selected an employee — resolve by legacy_id
        # For badge/manual: resolve by the scan value itself
        lookup_key = lookup_legacy_id if lookup_legacy_id else sanitized
        employee = self._employee_cache.get(lookup_key)

        # Check for duplicate badge scan (Issue #20)
        is_duplicate = False
        if config.DUPLICATE_BADGE_DETECTION_ENABLED:
            is_dup, original_id = self._db.check_if_duplicate_badge(
                sanitized,
                self.station_name,
                config.DUPLICATE_BADGE_TIME_WINDOW_SECONDS
            )
            is_duplicate = is_dup

            # If duplicate and action is 'block', reject the scan
            if is_duplicate and config.DUPLICATE_BADGE_ACTION == 'block':
                return {
                    "ok": False,
                    "status": "duplicate_rejected",
                    "message": f"Duplicate: Badge {sanitized} scanned within {config.DUPLICATE_BADGE_TIME_WINDOW_SECONDS} seconds",
                    "is_duplicate": True,
                    "badgeId": sanitized,
                    "fullName": employee.full_name if employee else "Unknown",
                }
        timestamp = datetime.now(timezone.utc).strftime(ISO_TIMESTAMP_FORMAT)
        self._db.record_scan(sanitized, self.station_name, employee, timestamp, scan_source=scan_source)
        history = self._db.get_recent_scans()
        # Only flag as duplicate for UI alert if action is 'warn' (not 'silent')
        # 'silent' mode accepts duplicates without any UI alert
        show_duplicate_alert = is_duplicate and config.DUPLICATE_BADGE_ACTION == 'warn'

        payload = {
            "ok": True,
            "badgeId": sanitized,
            "fullName": employee.full_name if employee else "Unknown",
            "matched": employee is not None,
            "timestamp": timestamp,
            "totalScansToday": self._db.count_scans_today(),
            "totalScansOverall": self._db.count_scans_total(),
            "scanHistory": [_scan_to_dict(scan) for scan in history],
            "is_duplicate": show_duplicate_alert,  # Only true for 'warn' mode
        }
        return payload

    def export_scans(self) -> Dict[str, object]:
        scans = self._db.fetch_all_scans()
        if not scans:
            return {
                "ok": False,
                "noData": True,
                "message": "No scan data to export.",
                "records": 0,
            }
        export_path = self._build_export_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scans"

        export_headers = [
            "Scan Value", "Legacy ID", "Full Name",
            "SL L1 Desc", "Position Desc", "Email",
            "Station", "Scanned At", "Matched", "Scan Source",
        ]
        sheet.append(export_headers)

        for record in scans:
            matched = record.legacy_id is not None
            row = [
                record.badge_id or "",
                record.legacy_id or "",
                record.employee_full_name or "",
                record.sl_l1_desc or "",
                record.position_desc or "",
                record.email or "",
                record.station_name or "",
                _format_timestamp(record.scanned_at),
                "Yes" if matched else "No",
                record.scan_source or "manual",
            ]
            sheet.append(row)

        for col_idx, header in enumerate(export_headers, start=1):
            max_length = len(header)
            for column_cells in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=sheet.max_row, values_only=True):
                for value in column_cells:
                    if value is None:
                        continue
                    max_length = max(max_length, len(str(value)))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)
        workbook.save(export_path)
        workbook.close()
        return {
            "ok": True,
            "fileName": export_path.name,
            "absolutePath": str(export_path),
            "records": len(scans),
        }

    def _build_export_path(self) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_station = _sanitize_filename_component(self.station_name)
        filename = f"Checkins_{safe_station}_{timestamp}.xlsx"
        return self._export_directory / filename

    def close(self) -> None:
        self._db.close()


def _sanitize_filename_component(component: str) -> str:
    fallback = "station"
    sanitized = re.sub(r"[^A-Za-z0-9_-]+", "-", component).strip("-")
    return sanitized or fallback


def _safe_string(value: Optional[object]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_timestamp(value: Optional[str]) -> str:
    """Convert UTC timestamp to local time for display."""
    if not value:
        return ""
    try:
        # Handle 'Z' suffix (UTC indicator) - replace with +00:00 for fromisoformat
        iso_value = value.replace('Z', '+00:00')
        utc_dt = datetime.fromisoformat(iso_value)
        # Convert to local timezone
        local_dt = utc_dt.astimezone()
        return local_dt.strftime(DISPLAY_TIMESTAMP_FORMAT)
    except ValueError:
        return value


def _scan_to_dict(record: ScanRecord) -> Dict[str, object]:
    return {
        "badgeId": record.badge_id,
        "timestamp": record.scanned_at,
        "fullName": record.employee_full_name or "Unknown",
        "station": record.station_name,
        "legacyId": record.legacy_id,
        "slL1Desc": record.sl_l1_desc,
        "positionDesc": record.position_desc,
        "matched": record.legacy_id is not None,
    }


__all__ = ["AttendanceService"]
